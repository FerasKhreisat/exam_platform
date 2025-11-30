from flask import Flask, render_template, request, redirect, url_for, session
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
from config import Config
from sqlalchemy import func
from werkzeug.utils import secure_filename
import os
app = Flask(__name__)
app.config.from_object(Config)
db = SQLAlchemy(app)

# ====================
# إعداد حساب الأدمن (من ملف config ليكون جاهز للاستضافة)
# ====================
ADMIN_EMAIL = Config.ADMIN_EMAIL
ADMIN_PASSWORD = Config.ADMIN_PASSWORD

# ====================
# ثوابت الصفوف / المسارات / الفصول
# ====================

# الصف الذي يُخزَّن مع الطالب (معلومات تعريفية فقط)
STUDENT_GRADES = [
    "العاشر",
    "الأول الثانوي",
    "الثاني الثانوي",
]

# لاختيار المواد داخل المنصة
LEVEL_CHOICES = STUDENT_GRADES  # نفس الأسماء
TRACK_CHOICES = ["أكاديمي", "مهني"]
SEMESTER_CHOICES = ["الفصل الأول", "الفصل الثاني"]


def build_grade_key(level: str, track: str, semester: str) -> str:
    """
    يبني مفتاح الصف للمادة، مثال:
    "الثاني الثانوي – أكاديمي – الفصل الأول"
    هذا النص هو الذي يجب أن تضعه في حقل grade عند إضافة المادة من الأدمن.
    """
    return f"{level} – {track} – {semester}"


# ====================
# نماذج قاعدة البيانات
# ====================

class Student(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    full_name = db.Column(db.String(120), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password = db.Column(db.String(120), nullable=False)
    # هنا نخزن الصف العام فقط: (العاشر / الأول الثانوي / الثاني الثانوي)
    grade = db.Column(db.String(50), nullable=False)
    results = db.relationship("ExamResult", backref="student", lazy=True)


class Subject(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    # هنا نخزن التركيبة الكاملة:
    # مثال: "الثاني الثانوي – أكاديمي – الفصل الأول"
    grade = db.Column(db.String(80), nullable=False)
    default_duration = db.Column(db.Integer, default=40)
    default_questions = db.Column(db.Integer, default=40)
    questions = db.relationship("Question", backref="subject", lazy=True)


class Question(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    subject_id = db.Column(db.Integer, db.ForeignKey("subject.id"), nullable=False)
    text = db.Column(db.Text, nullable=False)
    option1 = db.Column(db.String(255), nullable=False)
    option2 = db.Column(db.String(255), nullable=False)
    option3 = db.Column(db.String(255), nullable=False)
    option4 = db.Column(db.String(255), nullable=False)
    # 1 أو 2 أو 3 أو 4
    correct_option = db.Column(db.String(10), nullable=False)


class ExamResult(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey("student.id"), nullable=False)
    subject_id = db.Column(db.Integer, db.ForeignKey("subject.id"), nullable=False)

    # علاقة بالمادة لسهولة الوصول للاسم
    subject = db.relationship("Subject", backref="results")

    score = db.Column(db.Float, nullable=False)
    correct_count = db.Column(db.Integer, nullable=False)
    wrong_count = db.Column(db.Integer, nullable=False)
    date = db.Column(db.DateTime, default=datetime.utcnow)
    answers = db.relationship("ExamAnswer", backref="result", lazy=True)


class ExamAnswer(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    result_id = db.Column(db.Integer, db.ForeignKey("exam_result.id"), nullable=False)
    question_id = db.Column(db.Integer, db.ForeignKey("question.id"), nullable=False)
    student_answer = db.Column(db.String(10), nullable=False)
    is_correct = db.Column(db.Boolean, nullable=False)

# إنشاء الجداول مرة واحدة عند بدء التطبيق (لـ Render أو التشغيل العادي)
with app.app_context():
    db.create_all()
    
# ====================
#  مسارات الطلاب
# ====================

@app.route("/")
def index():
    return redirect(url_for("login"))


@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        full_name = request.form.get("full_name")
        email = request.form.get("email")
        grade = request.form.get("grade")  # واحد من STUDENT_GRADES
        password = request.form.get("password")

        # تحقق من صحة الصف
        if grade not in STUDENT_GRADES:
            return render_template(
                "register.html",
                grades=STUDENT_GRADES,
                error="الصف المختار غير صحيح."
            )

        existing = Student.query.filter_by(email=email).first()
        if existing:
            return render_template(
                "register.html",
                grades=STUDENT_GRADES,
                error="البريد الإلكتروني مستخدم مسبقاً!"
            )

        student = Student(
            full_name=full_name,
            email=email,
            grade=grade,
            password=password
        )
        db.session.add(student)
        db.session.commit()

        return redirect(url_for("login"))

    # GET
    return render_template("register.html", grades=STUDENT_GRADES)


@app.route("/forgot_password", methods=["GET", "POST"])
def forgot_password():
    if request.method == "POST":
        email = request.form.get("email")

        user = Student.query.filter_by(email=email).first()
        if not user:
            return render_template("forgot_password.html", error="البريد غير موجود!")

        # حفظ البريد في جلسة مؤقتة
        session["reset_email"] = email
        return redirect(url_for("reset_password"))

    return render_template("forgot_password.html")


@app.route("/reset_password", methods=["GET", "POST"])
def reset_password():
    if "reset_email" not in session:
        return redirect(url_for("forgot_password"))

    if request.method == "POST":
        new_password = request.form.get("password")

        user = Student.query.filter_by(email=session["reset_email"]).first()
        if user:
            user.password = new_password
            db.session.commit()

        session.pop("reset_email", None)
        return redirect(url_for("login"))

    return render_template("reset_password.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email")
        password = request.form.get("password")

        student = Student.query.filter_by(email=email, password=password).first()

        if student:
            session["student_id"] = student.id
            session["student_name"] = student.full_name
            session["grade"] = student.grade
            # عند أول دخول لا يوجد اختيار فصل/مسار بعد
            session.pop("current_grade_key", None)
            return redirect(url_for("choose_grade"))
        return render_template("login.html", error="بيانات الدخول غير صحيحة!")

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/student/dashboard")
def student_dashboard():
    """لوحة الطالب التي تُظهر السجل والإحصائيات."""
    if "student_id" not in session:
        return redirect(url_for("login"))

    student = Student.query.get_or_404(session["student_id"])

    # كل نتائج الطالب مع اسم المادة
    results_query = (
        db.session.query(ExamResult, Subject.name.label("subject_name"))
        .join(Subject, ExamResult.subject_id == Subject.id)
        .filter(ExamResult.student_id == student.id)
        .order_by(ExamResult.date.desc())
        .all()
    )

    results = []
    total_correct = 0
    total_wrong = 0

    for res, subject_name in results_query:
        results.append({
            "subject_name": subject_name,
            "date": res.date.strftime("%Y-%m-%d %H:%M"),
            "score": res.score,
            "correct": res.correct_count,
            "wrong": res.wrong_count,
        })
        total_correct += res.correct_count
        total_wrong += res.wrong_count

    total_exams = len(results)

    return render_template(
        "student_dashboard.html",
        student=student,
        total_exams=total_exams,
        total_correct=total_correct,
        total_wrong=total_wrong,
        results=results,
    )
@app.route("/contact")
def contact():
    return render_template("contact.html")

# ====================
# اختيار الصف / المسار / الفصل داخل المنصة
# ====================

@app.route("/choose_grade", methods=["GET", "POST"])
def choose_grade():
    if "student_id" not in session:
        return redirect(url_for("login"))

    if request.method == "POST":
        level = request.form.get("level")
        track = request.form.get("track")
        semester = request.form.get("semester")

        if not level or not track or not semester:
            return render_template(
                "choose_grade.html",
                levels=LEVEL_CHOICES,
                tracks=TRACK_CHOICES,
                semesters=SEMESTER_CHOICES,
                error="يرجى اختيار الصف والمسار والفصل."
            )

        grade_key = build_grade_key(level, track, semester)
        session["current_grade_key"] = grade_key

        return redirect(url_for("student_subjects"))

    return render_template(
        "choose_grade.html",
        levels=LEVEL_CHOICES,
        tracks=TRACK_CHOICES,
        semesters=SEMESTER_CHOICES
    )


@app.route("/student/subjects")
def student_subjects():
    if "student_id" not in session:
        return redirect(url_for("login"))

    grade_key = session.get("current_grade_key")
    if not grade_key:
        # لم يختَر الصف بعد
        return redirect(url_for("choose_grade"))

    subjects = Subject.query.filter_by(grade=grade_key).order_by(Subject.name).all()

    return render_template("student_subjects.html", subjects=subjects)

# ====================
# الامتحان
# ====================

@app.route("/exam/start/<int:subject_id>", methods=["GET", "POST"])
def exam_start(subject_id):
    if "student_id" not in session:
        return redirect(url_for("login"))

    subject = Subject.query.get_or_404(subject_id)

    if request.method == "POST":
        total_q = subject.default_questions
        questions = (
            Question.query
            .filter_by(subject_id=subject.id)
            .order_by(func.random())
            .limit(total_q)
            .all()
        )

        if not questions:
            return render_template(
                "exam_start.html",
                subject=subject,
                error="لا توجد أسئلة لهذه المادة بعد."
            )

        # إنشاء سجل نتيجة مبدئي
        result = ExamResult(
            student_id=session["student_id"],
            subject_id=subject.id,
            score=0,
            correct_count=0,
            wrong_count=0,
        )
        db.session.add(result)
        db.session.commit()

        # حفظ بيانات الامتحان في الـ session
        session["exam_id"] = result.id
        session["exam_questions"] = [q.id for q in questions]
        session["exam_index"] = 0
        session["exam_duration"] = subject.default_duration  # بالدقائق

        return redirect(url_for("exam_take"))

    return render_template("exam_start.html", subject=subject)


def _finalize_exam(exam_id, q_ids):
    """حساب النتيجة النهائية وتنظيف الـ session."""
    result = ExamResult.query.get_or_404(exam_id)
    total = len(q_ids) if q_ids else 0

    # لو ما تم احتساب عدد الأسئلة الخاطئة نكمله هنا
    if total and result.correct_count + result.wrong_count != total:
        result.wrong_count = total - result.correct_count

    if total:
        result.score = (result.correct_count / total) * 100
    else:
        result.score = 0

    db.session.commit()

    # تنظيف بيانات الامتحان من الجلسة
    session.pop("exam_id", None)
    session.pop("exam_questions", None)
    session.pop("exam_index", None)
    session.pop("exam_duration", None)

    return result


@app.route("/exam/take", methods=["GET", "POST"])
def exam_take():
    if "student_id" not in session:
        return redirect(url_for("login"))

    exam_id = session.get("exam_id")
    q_ids = session.get("exam_questions")

    if not exam_id or not q_ids:
        return redirect(url_for("student_subjects"))

    # رقم السؤال الحالي المخزَّن في السيشن
    index = session.get("exam_index", 0)

    # لو جاء رقم سؤال من شريط التنقّل (GET ?index=...)
    if request.method == "GET":
        idx_from_url = request.args.get("index")
        if idx_from_url is not None:
            try:
                idx_from_url = int(idx_from_url)
            except ValueError:
                idx_from_url = 0

            # منع الخروج عن حدود الأسئلة
            idx_from_url = max(0, min(idx_from_url, len(q_ids) - 1))
            index = idx_from_url
            session["exam_index"] = index

    # حفظ إجابة الطالب عند الضغط على "التالي" أو "إنهاء الامتحان"
    if request.method == "POST":
        selected = request.form.get("answer")
        question_id = int(request.form.get("question_id"))
        current_index = int(request.form.get("current_index", index))
        action = request.form.get("action", "next")

        result = ExamResult.query.get_or_404(exam_id)
        question = Question.query.get_or_404(question_id)

        if selected:
            # هل يوجد إجابة سابقة لنفس السؤال؟
            answer_row = ExamAnswer.query.filter_by(
                result_id=exam_id,
                question_id=question.id
            ).first()

            is_correct = (selected == question.correct_option)

            if answer_row:
                # عدّل عدّاد الصح/الخطأ لو تغيّرت النتيجة
                if answer_row.is_correct and not is_correct:
                    result.correct_count -= 1
                    result.wrong_count += 1
                elif not answer_row.is_correct and is_correct:
                    result.correct_count += 1
                    result.wrong_count -= 1

                answer_row.student_answer = selected
                answer_row.is_correct = is_correct
            else:
                answer_row = ExamAnswer(
                    result_id=exam_id,
                    question_id=question.id,
                    student_answer=selected,
                    is_correct=is_correct
                )
                db.session.add(answer_row)

                if is_correct:
                    result.correct_count += 1
                else:
                    result.wrong_count += 1

            db.session.commit()

        # لو ضغط "إنهاء الامتحان" ننهي مباشرة
        if action == "finish":
            total = len(q_ids) or 1
            result.score = (result.correct_count / total) * 100
            db.session.commit()

            # تنظيف بيانات الامتحان من السيشن
            session.pop("exam_id", None)
            session.pop("exam_questions", None)
            session.pop("exam_index", None)
            session.pop("exam_duration", None)

            return redirect(url_for("exam_result", result_id=result.id))

        # الانتقال للسؤال التالي
        index = current_index + 1
        session["exam_index"] = index

        # لو وصلنا لنهاية الأسئلة نحسب النتيجة وننهي
        if index >= len(q_ids):
            total = len(q_ids) or 1
            result.score = (result.correct_count / total) * 100
            db.session.commit()

            session.pop("exam_id", None)
            session.pop("exam_questions", None)
            session.pop("exam_index", None)
            session.pop("exam_duration", None)

            return redirect(url_for("exam_result", result_id=result.id))

    # تأمين عدم الخروج عن النطاق
    if index >= len(q_ids):
        index = len(q_ids) - 1

    current_q_id = q_ids[index]
    question = Question.query.get_or_404(current_q_id)

    duration = session.get("exam_duration", 40)
    total_questions = len(q_ids)

    # تجهيز بيانات الإجابات لاستخدامها في الواجهة
    all_answers = ExamAnswer.query.filter_by(result_id=exam_id).all()
    answers_by_qid = {a.question_id: a for a in all_answers}

    # الإجابة المحفوظة للسؤال الحالي (لـ saved_answer في الـ HTML)
    current_answer_row = answers_by_qid.get(current_q_id)
    saved_answer = None
    if current_answer_row and current_answer_row.student_answer:
        try:
            saved_answer = int(current_answer_row.student_answer)
        except ValueError:
            saved_answer = None

    # الأسئلة التي تمّت الإجابة عنها (لـ answered في شريط الأرقام)
    answered = {}
    for i, qid in enumerate(q_ids):
        if qid in answers_by_qid:
            answered[i] = True

    return render_template(
        "exam_take.html",
        question=question,
        index=index,
        total_questions=total_questions,
        duration=duration,
        saved_answer=saved_answer,
        answered=answered,
    )


@app.route("/exam/result/<int:result_id>")
def exam_result(result_id):
    if "student_id" not in session:
        return redirect(url_for("login"))

    result = ExamResult.query.get_or_404(result_id)

    # منع طالب آخر من رؤية نتيجة غيره
    if result.student_id != session["student_id"]:
        return redirect(url_for("student_subjects"))

    answers = ExamAnswer.query.filter_by(result_id=result.id).all()

    # تجهيز بيانات الأسئلة مع الإجابات
    detailed = []
    for ans in answers:
        q = Question.query.get(ans.question_id)
        detailed.append({
            "text": q.text,
            "option1": q.option1,
            "option2": q.option2,
            "option3": q.option3,
            "option4": q.option4,
            "correct_option": q.correct_option,
            "student_answer": ans.student_answer,
            "is_correct": ans.is_correct,
        })

    return render_template(
        "exam_result.html",
        result=result,
        answers=detailed,
    )


@app.route("/exam/result/<int:result_id>/pdf")
def exam_result_pdf(result_id):
    """عرض تقرير النتيجة كصفحة HTML جاهزة للطباعة."""
    if "student_id" not in session:
        return redirect(url_for("login"))

    result = ExamResult.query.get_or_404(result_id)

    if result.student_id != session["student_id"]:
        return redirect(url_for("student_subjects"))

    answers = ExamAnswer.query.filter_by(result_id=result.id).all()

    detailed = []
    for ans in answers:
        q = Question.query.get(ans.question_id)
        detailed.append({
            "text": q.text,
            "option1": q.option1,
            "option2": q.option2,
            "option3": q.option3,
            "option4": q.option4,
            "correct_option": q.correct_option,
            "student_answer": ans.student_answer,
            "is_correct": ans.is_correct,
        })

    return render_template(
        "result_pdf.html",
        result=result,
        answers=detailed,
    )



@app.route("/exam/finish")
def exam_finish():
    """يُستدعى من عدّاد الوقت عند انتهاء الزمن."""
    if "student_id" not in session:
        return redirect(url_for("login"))

    exam_id = session.get("exam_id")
    q_ids = session.get("exam_questions")

    if not exam_id or not q_ids:
        return redirect(url_for("student_subjects"))

    final_result = _finalize_exam(exam_id, q_ids)
    return redirect(url_for("exam_result", result_id=final_result.id))


# ====================
#   مسارات الأدمن
# ====================

def admin_required():
    return session.get("is_admin")


def get_subject_grades():
    """
    يبني قائمة بكل التركيبات الممكنة للصف / المسار / الفصل
    مثل: 'الثاني الثانوي – أكاديمي – الفصل الأول'
    وتُستخدم في صفحات إضافة/تعديل المادة.
    """
    grades = []
    for level in LEVEL_CHOICES:
        for track in TRACK_CHOICES:
            for semester in SEMESTER_CHOICES:
                grades.append(build_grade_key(level, track, semester))
    return grades


@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        email = request.form.get("email")
        password = request.form.get("password")

        if email == ADMIN_EMAIL and password == ADMIN_PASSWORD:
            session["is_admin"] = True
            return redirect(url_for("admin_dashboard"))

        return render_template("admin_login.html", error="خطأ في بيانات الدخول!")

    return render_template("admin_login.html")


@app.route("/admin/logout")
def admin_logout():
    session.pop("is_admin", None)
    return redirect(url_for("admin_login"))


@app.route("/admin/dashboard")
def admin_dashboard():
    if not admin_required():
        return redirect(url_for("admin_login"))

    stats = {
        "students": Student.query.count(),
        "subjects": Subject.query.count(),
        "questions": Question.query.count(),
        "results": ExamResult.query.count(),
    }

    recent = (
        db.session.query(
            ExamResult.id,
            Student.full_name.label("student_name"),
            Subject.name.label("subject_name"),
            ExamResult.score,
            ExamResult.date,
        )
        .join(Student, ExamResult.student_id == Student.id)
        .join(Subject, ExamResult.subject_id == Subject.id)
        .order_by(ExamResult.date.desc())
        .limit(5)
        .all()
    )

    return render_template("admin_dashboard.html", stats=stats, recent=recent)


# ---------- المواد ----------

@app.route("/admin/subjects")
def admin_subjects():
    if not admin_required():
        return redirect(url_for("admin_login"))

    subjects = Subject.query.order_by(Subject.grade, Subject.name).all()
    return render_template("admin_subjects.html", subjects=subjects)


@app.route("/admin/subjects/add", methods=["GET", "POST"])
def admin_add_subject():
    if not admin_required():
        return redirect(url_for("admin_login"))

    grades = get_subject_grades()
    error = None

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        grade = request.form.get("grade")
        duration = request.form.get("duration")
        questions = request.form.get("questions")

        if not name or not grade or not duration or not questions:
            error = "يرجى تعبئة جميع الحقول."
        elif grade not in grades:
            error = "الصف المختار غير صحيح."
        else:
            try:
                duration = int(duration)
                questions = int(questions)
                if duration <= 0 or questions <= 0:
                    error = "يجب أن تكون المدة وعدد الأسئلة أرقامًا موجبة."
            except ValueError:
                error = "مدة الامتحان وعدد الأسئلة يجب أن تكون أرقامًا."

        if not error:
            subject = Subject(
                name=name,
                grade=grade,
                default_duration=duration,
                default_questions=questions
            )
            db.session.add(subject)
            db.session.commit()
            return redirect(url_for("admin_subjects"))

    return render_template(
        "admin_add_subject.html",
        grades=grades,
        error=error
    )


@app.route("/admin/subjects/<int:subject_id>/edit", methods=["GET", "POST"])
def admin_edit_subject(subject_id):
    if not admin_required():
        return redirect(url_for("admin_login"))

    subject = Subject.query.get_or_404(subject_id)
    grades = get_subject_grades()
    error = None

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        grade = request.form.get("grade")
        duration = request.form.get("duration")
        questions = request.form.get("questions")

        if not name or not grade or not duration or not questions:
            error = "يرجى تعبئة جميع الحقول."
        elif grade not in grades:
            error = "الصف المختار غير صحيح."
        else:
            try:
                duration = int(duration)
                questions = int(questions)
                if duration <= 0 or questions <= 0:
                    error = "يجب أن تكون المدة وعدد الأسئلة أرقامًا موجبة."
            except ValueError:
                error = "مدة الامتحان وعدد الأسئلة يجب أن تكون أرقامًا."

        if not error:
            subject.name = name
            subject.grade = grade
            subject.default_duration = duration
            subject.default_questions = questions
            db.session.commit()
            return redirect(url_for("admin_subjects"))

    return render_template(
        "admin_edit_subject.html",
        subject=subject,
        grades=grades,
        error=error
    )


@app.route("/admin/subjects/<int:subject_id>/delete")
def admin_delete_subject(subject_id):
    if not admin_required():
        return redirect(url_for("admin_login"))

    subject = Subject.query.get_or_404(subject_id)

    # حذف النتائج المرتبطة بالمادة + الإجابات
    results = ExamResult.query.filter_by(subject_id=subject.id).all()
    for r in results:
        ExamAnswer.query.filter_by(result_id=r.id).delete()
        db.session.delete(r)

    # حذف الأسئلة + إجاباتها
    questions = Question.query.filter_by(subject_id=subject.id).all()
    for q in questions:
        ExamAnswer.query.filter_by(question_id=q.id).delete()
        db.session.delete(q)

    db.session.delete(subject)
    db.session.commit()

    return redirect(url_for("admin_subjects"))


# ---------- الأسئلة ----------

@app.route("/admin/questions/<int:subject_id>", methods=["GET", "POST"])
def admin_questions(subject_id):
    if not admin_required():
        return redirect(url_for("admin_login"))

    subject = Subject.query.get_or_404(subject_id)

    if request.method == "POST":
        text = request.form.get("text")
        option1 = request.form.get("option1")
        option2 = request.form.get("option2")
        option3 = request.form.get("option3")
        option4 = request.form.get("option4")
        correct_option = request.form.get("correct_option")  # 1/2/3/4

        if text and option1 and option2 and option3 and option4 and correct_option:
            question = Question(
                subject_id=subject.id,
                text=text,
                option1=option1,
                option2=option2,
                option3=option3,
                option4=option4,
                correct_option=correct_option
            )
            db.session.add(question)
            db.session.commit()

    questions = Question.query.filter_by(subject_id=subject.id).all()
    return render_template(
        "admin_questions.html",
        subject=subject,
        questions=questions
    )


@app.route("/admin/questions/detail/<int:question_id>")
def admin_question_detail(question_id):
    if not admin_required():
        return redirect(url_for("admin_login"))

    question = Question.query.get_or_404(question_id)
    return render_template("admin_question_detail.html", question=question)


@app.route("/admin/questions/<int:question_id>/edit", methods=["GET", "POST"])
def admin_edit_question(question_id):
    if not admin_required():
        return redirect(url_for("admin_login"))

    question = Question.query.get_or_404(question_id)

    if request.method == "POST":
        question.text = request.form.get("text")
        question.option1 = request.form.get("option1")
        question.option2 = request.form.get("option2")
        question.option3 = request.form.get("option3")
        question.option4 = request.form.get("option4")
        question.correct_option = request.form.get("correct_option")
        db.session.commit()
        return redirect(url_for("admin_question_detail", question_id=question.id))

    return render_template("admin_edit_question.html", question=question)


@app.route("/admin/questions/<int:question_id>/delete")
def admin_delete_question(question_id):
    if not admin_required():
        return redirect(url_for("admin_login"))

    question = Question.query.get_or_404(question_id)
    subject_id = question.subject_id

    ExamAnswer.query.filter_by(question_id=question.id).delete()
    db.session.delete(question)
    db.session.commit()

    return redirect(url_for("admin_questions", subject_id=subject_id))


# ---------- رفع ملف أسئلة (إكسل) ----------

@app.route("/admin/upload", methods=["GET", "POST"])
def admin_upload():
    if not admin_required():
        return redirect(url_for("admin_login"))

    subjects = Subject.query.order_by(Subject.grade, Subject.name).all()
    message = None
    error = None

    if request.method == "POST":
        subject_id = request.form.get("subject_id")
        file = request.files.get("file")

        if not subject_id or not file:
            error = "يرجى اختيار المادة ورفع الملف."
        else:
            try:
                filename = secure_filename(file.filename)
                upload_dir = os.path.join(app.root_path, "uploads")
                os.makedirs(upload_dir, exist_ok=True)
                filepath = os.path.join(upload_dir, filename)
                file.save(filepath)

                # قراءة ملف إكسل
                import openpyxl
                wb = openpyxl.load_workbook(filepath)
                sheet = wb.active

                # نقرأ صف العناوين (الهيدر)
                header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
                headers = [(str(h).strip().lower() if h else "") for h in header_row]

                def get_index(col_name: str) -> int:
                    """إرجاع رقم العمود حسب اسم الهيدر، أو رفع خطأ إن لم يوجد."""
                    try:
                        return headers.index(col_name.lower())
                    except ValueError:
                        raise ValueError(f"لم يتم العثور على العمود '{col_name}' في ملف الإكسل")

                # تحديد أعمدة السؤال والخيارات والإجابة الصحيحة
                q_idx       = get_index("question")
                opt_a_idx   = get_index("option_a")
                opt_b_idx   = get_index("option_b")
                opt_c_idx   = get_index("option_c")
                opt_d_idx   = get_index("option_d")
                correct_idx = get_index("correct")

                added = 0
                # نبدأ من السطر الثاني (بعد الهيدر)
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    # تخطّي الصفوف الفارغة أو التي لا تحتوي سؤالاً
                    if not row or q_idx >= len(row) or not row[q_idx]:
                        continue

                    text    = str(row[q_idx])
                    o1      = str(row[opt_a_idx]) if opt_a_idx < len(row) else ""
                    o2      = str(row[opt_b_idx]) if opt_b_idx < len(row) else ""
                    o3      = str(row[opt_c_idx]) if opt_c_idx < len(row) else ""
                    o4      = str(row[opt_d_idx]) if opt_d_idx < len(row) else ""
                    correct = str(row[correct_idx]).strip() if correct_idx < len(row) else ""

                    q = Question(
                        subject_id=int(subject_id),
                        text=text,
                        option1=o1,
                        option2=o2,
                        option3=o3,
                        option4=o4,
                        correct_option=correct,
                    )
                    db.session.add(q)
                    added += 1

                db.session.commit()
                message = f"تم استيراد {added} سؤالاً بنجاح."

            except Exception:
                db.session.rollback()
                error = "حدث خطأ أثناء قراءة الملف. تأكد من أن التنسيق صحيح وأن عناوين الأعمدة مكتوبة بشكل صحيح."

    return render_template(
        "admin_upload.html",
        subjects=subjects,
        message=message,
        error=error
    )


# ---------- الطلاب ----------

@app.route("/admin/students")
def admin_students():
    if not admin_required():
        return redirect(url_for("admin_login"))

    students = Student.query.order_by(Student.full_name).all()
    return render_template("admin_students.html", students=students)


@app.route("/admin/students/<int:student_id>/delete")
def admin_delete_student(student_id):
    if not admin_required():
        return redirect(url_for("admin_login"))

    student = Student.query.get_or_404(student_id)

    results = ExamResult.query.filter_by(student_id=student.id).all()
    for r in results:
        ExamAnswer.query.filter_by(result_id=r.id).delete()
        db.session.delete(r)

    db.session.delete(student)
    db.session.commit()
    return redirect(url_for("admin_students"))


# ---------- النتائج ----------

@app.route("/admin/results")
def admin_results():
    if not admin_required():
        return redirect(url_for("admin_login"))

    results = (
        db.session.query(
            ExamResult.id,
            Student.full_name.label("student_name"),
            Subject.name.label("subject_name"),
            ExamResult.score,
            ExamResult.date,
        )
        .join(Student, ExamResult.student_id == Student.id)
        .join(Subject, ExamResult.subject_id == Subject.id)
        .order_by(ExamResult.date.desc())
        .all()
    )

    return render_template("admin_results.html", results=results)


@app.route("/admin/results/<int:result_id>/delete")
def admin_delete_result(result_id):
    if not admin_required():
        return redirect(url_for("admin_login"))

    result = ExamResult.query.get_or_404(result_id)
    ExamAnswer.query.filter_by(result_id=result.id).delete()
    db.session.delete(result)
    db.session.commit()

    return redirect(url_for("admin_results"))


# ====================
# تشغيل السيرفر (محلياً فقط)
# ====================
if __name__ == "__main__":
    app.run(debug=True)